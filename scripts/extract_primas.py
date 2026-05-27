"""Extrae las tarifas INS Medical del XLSX a un JS module.

Estructura del XLSX (3 hojas: REGIONAL, INTERNACIONAL, GRANDES DEDUCIBLES):
- Cada hoja tiene 2 secciones: "CON REDUCCIÓN DE SUMA ASEGURADA" y "SIN REDUCCIÓN".
- Cada sección: encabezado en row N, planes en row N+1, género en row N+2, datos desde row N+3.
- REGIONAL e INTERNACIONAL: 6 planes (Sin ded, $200, $300, $400, $500, $1000).
- GRANDES DEDUCIBLES: 3 planes ($5000, $10000, $15000).
- Columnas pares = Femenino, impares = Masculino (después de la columna A=Edad).
- Edades 0 a 99 (99 = 99+).

Output: scripts/primas_dump.json + scripts/primas_data.js
"""
import json
from pathlib import Path
import openpyxl

XLSX_PATH_BASE = Path(r"C:\Users\segur\Ins Medical\PRIMAS 2025 Dic 25.xlsx")             # sin AMP
XLSX_PATH_AMP  = Path(r"C:\Users\segur\Ins Medical\PRIMAS 2025 AMP EPID Y PAND Dic 25.xlsx")  # con AMP
OUT_JSON = Path(r"C:\Users\segur\cotizador-Ins-Medical\scripts\primas_dump.json")
OUT_JS = Path(r"C:\Users\segur\cotizador-Ins-Medical\scripts\primas_data.js")

# Mapeo de etiqueta de columna del XLSX a clave de deducible en JS.
DEDUCIBLES_REG_INT = {
    "Regional sin deducible (Prórroga)": "0",
    "Regional $200 Deducible": "200",
    "Regional $300 Deducible": "300",
    "Regional $400 Deducible": "400",
    "Regional $500 Deducible": "500",
    "Regional $1000 Deducible": "1000",
    "Internacional sin deducible (Prórroga)": "0",
    "Internacional $200 Deducible": "200",
    "Internacional $300 Deducible": "300",
    "Internacional $400 Deducible": "400",
    "Internacional $500 Deducible": "500",
    "Internacional $1000 Deducible": "1000",
}
DEDUCIBLES_GD = {
    "GRANDES DEDUCIBLES $5000": "5000",
    "GRANDES DEDUCIBLES $10000": "10000",
    "GRANDES DEDUCIBLES $15000": "15000",
    # Variantes posibles del título
    "Grandes Deducibles $5000": "5000",
    "Grandes Deducibles $10000": "10000",
    "Grandes Deducibles $15000": "15000",
}


def _norm(s):
    if s is None:
        return ""
    return str(s).strip().replace("\xa0", " ").replace("  ", " ")


def find_sections(ws):
    """Devuelve [(label, start_row), ...] con CON/SIN REDUCCIÓN."""
    out = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = _norm(cell.value).upper()
            if "REDUCCI" in v and "SUMA ASEGURADA" in v:
                out.append((v, cell.row))
    return out


def parse_section(ws, header_row, expected_deductibles):
    """Parsea una sección a partir de la fila del header 'CON/SIN REDUCCIÓN'.

    Estructura: header_row es la fila del título. Las siguientes filas son:
      header_row + 1: nombre del producto (e.g. 'REGIONAL' / 'INTERNACIONAL')
      header_row + 2: nombres de planes en columnas B/D/F/...
      header_row + 3: 'Femenino' / 'Masculino' alternándose
      header_row + 4 en adelante: datos edad 0..99
    """
    # Detectar la fila de los planes (la primera con "deducible" o "GRANDES")
    plan_row = None
    for offset in range(1, 6):
        r = header_row + offset
        for col in range(2, ws.max_column + 1):
            v = _norm(ws.cell(row=r, column=col).value)
            if v and ("deducible" in v.lower() or "GRANDES" in v.upper() or "Prórroga" in v):
                plan_row = r
                break
        if plan_row:
            break
    assert plan_row is not None, f"No encontré fila de planes para sección @{header_row}"

    # Mapear cada columna a un deducible.
    # Los planes están en columnas pares (B, D, F, ...) por mergedCells.
    col_to_ded = {}  # col -> (deducible_key, gender)
    last_label = None
    for col in range(2, ws.max_column + 1):
        v = _norm(ws.cell(row=plan_row, column=col).value)
        if v and v in expected_deductibles:
            last_label = expected_deductibles[v]
        # Fila plan_row+1 dice gender
        gv = _norm(ws.cell(row=plan_row + 1, column=col).value).lower()
        gender = None
        if gv.startswith("fem"):
            gender = "F"
        elif gv.startswith("mas"):
            gender = "M"
        if last_label and gender:
            col_to_ded[col] = (last_label, gender)

    # Leer datos
    data_start = plan_row + 2
    rows_data = {}  # edad -> {ded: {F: prima, M: prima}}
    for r in range(data_start, ws.max_row + 1):
        age_cell = ws.cell(row=r, column=1).value
        if age_cell is None or age_cell == "":
            # Filas vacías o de separación - posiblemente fin de sección
            continue
        # Edad puede ser int o str
        try:
            if isinstance(age_cell, str):
                age_s = age_cell.strip().replace("+", "")
                age = int(age_s)
            else:
                age = int(age_cell)
        except Exception:
            # No es edad - probablemente otro encabezado, salimos.
            break
        if age < 0 or age > 99:
            break
        row_data = {}
        for col, (ded, gender) in col_to_ded.items():
            v = ws.cell(row=r, column=col).value
            if v is None or v == "":
                continue
            try:
                prima = int(v) if isinstance(v, (int, float)) else int(float(str(v).replace(",", "")))
            except Exception:
                continue
            row_data.setdefault(ded, {})[gender] = prima
        rows_data[age] = row_data
    return rows_data


def parse_sheet(ws, sheet_name, expected_deductibles):
    """Parsea una hoja entera devolviendo {'con_reduccion': {edad: {...}}, 'sin_reduccion': {edad: {...}}}."""
    sections = find_sections(ws)
    print(f"[{sheet_name}] secciones detectadas: {len(sections)}")
    for label, row in sections:
        print(f"    @row {row}: {label}")
    out = {}
    for label, row in sections:
        key = "con_reduccion" if "CON" in label.split("REDUCCI")[0].upper() else "sin_reduccion"
        # Más simple: si empieza con "CON" → con_reduccion, sino sin_reduccion
        key = "con_reduccion" if label.startswith("CON") else "sin_reduccion"
        out[key] = parse_section(ws, row, expected_deductibles)
        print(f"    sección {key}: {len(out[key])} edades parseadas")
    return out


def parse_workbook(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return {
        "regional":           parse_sheet(wb["REGIONAL"],           "REGIONAL",           DEDUCIBLES_REG_INT),
        "internacional":      parse_sheet(wb["INTERNACIONAL"],      "INTERNACIONAL",      DEDUCIBLES_REG_INT),
        "grandes_deducibles": parse_sheet(wb["GRANDES DEDUCIBLES"], "GRANDES DEDUCIBLES", DEDUCIBLES_GD),
    }


def main():
    print("=" * 80)
    print("PARSE BASE (sin AMP) →", XLSX_PATH_BASE.name)
    print("=" * 80)
    base = parse_workbook(XLSX_PATH_BASE)

    print("\n" + "=" * 80)
    print("PARSE AMP (con AMP) →", XLSX_PATH_AMP.name)
    print("=" * 80)
    amp = parse_workbook(XLSX_PATH_AMP)

    # Resultado: estructura combinada con dos versiones por producto.
    result = {
        "base": base,  # sin ampliación de epidemias y pandemias
        "amp":  amp    # con ampliación de epidemias y pandemias
    }

    # Validación: cada versión × producto × sección × edad 0..99 × deducible × F/M
    print("\n" + "=" * 80)
    print("VALIDACIÓN")
    print("=" * 80)
    expected_ded_reg = {"0", "200", "300", "400", "500", "1000"}
    expected_ded_gd = {"5000", "10000", "15000"}
    missing = []
    for ver_key, ver_data in result.items():
        for prod_key, prod_data in ver_data.items():
            ded_expected = expected_ded_gd if prod_key == "grandes_deducibles" else expected_ded_reg
            for sec_key, sec_data in prod_data.items():
                for age in range(0, 100):
                    if age not in sec_data:
                        missing.append(f"{ver_key}/{prod_key}/{sec_key}: edad {age} faltante")
                        continue
                    for ded in ded_expected:
                        if ded not in sec_data[age]:
                            missing.append(f"{ver_key}/{prod_key}/{sec_key}: edad {age} deducible {ded} faltante")
                            continue
                        for g in ("F", "M"):
                            if g not in sec_data[age][ded]:
                                missing.append(f"{ver_key}/{prod_key}/{sec_key}: edad {age} deducible {ded} género {g} faltante")
    if missing:
        print(f"FALLA: {len(missing)} faltantes (primeros 20):")
        for m in missing[:20]:
            print(f"  - {m}")
    else:
        print("OK - base+amp × 100 edades × deducibles × F/M completo en las 12 secciones (2v × 3p × 2s).")

    # Dump JSON
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nJSON: {OUT_JSON}")

    # Dump JS module
    js_lines = [
        "// Tarifas INS Medical - vigentes desde 01/12/2025",
        "// Fuente: PRIMAS 2025 AMP EPID Y PAND Dic 25.xlsx (INS oficial)",
        "// Generado automáticamente por scripts/extract_primas.py - NO EDITAR A MANO",
        "// Estructura: PRIMAS[producto][seccion][edad][deducible][genero] = prima_usd_anual",
        "// producto: 'regional' | 'internacional' | 'grandes_deducibles'",
        "// seccion: 'con_reduccion' | 'sin_reduccion'  (sin_reduccion se aplica desde edad 70 si JC elige)",
        "// edad: 0..99 (99 representa '99 o más')",
        "// deducible (REG/INT): '0'|'200'|'300'|'400'|'500'|'1000'   ;   (GD): '5000'|'10000'|'15000'",
        "// genero: 'F' | 'M'",
        "",
        "export const PRIMAS_VERSION = 'INS-MEDICAL-2025-12-01';",
        "",
        "export const PRIMAS = " + json.dumps(result, indent=2, ensure_ascii=False) + ";",
        "",
    ]
    OUT_JS.write_text("\n".join(js_lines), encoding="utf-8")
    print(f"JS:   {OUT_JS}")

    # Sanity check vs PDF oficial cotizador INS (Tierney) — Grandes Deducibles $5000 sin AMP, con IVA
    print("\n" + "=" * 80)
    print("SANITY CHECK contra PDF oficial INS (Tierney) - GD $5000 sin AMP")
    print("=" * 80)
    print("PDF dice: anual $1,129.14 (con IVA 2%). Sin IVA = $1,107.00")
    gd_base = result["base"]["grandes_deducibles"]["con_reduccion"]
    print("BASE (sin AMP) Grandes Ded $5000 con_red:")
    for age in range(18, 65, 5):
        try:
            f = gd_base[age]['5000']['F']
            m = gd_base[age]['5000']['M']
            iva_m = m * 1.02
            print(f"  edad {age:2d}: F={f:>5}  M={m:>5}  (M con IVA = ${iva_m:7.2f})")
        except KeyError as e:
            print(f"  edad {age}: FALTA {e}")
    # AMP comparison
    print("\nAMP (con AMP) Grandes Ded $5000 con_red (comparación):")
    gd_amp = result["amp"]["grandes_deducibles"]["con_reduccion"]
    for age in range(18, 65, 5):
        try:
            f = gd_amp[age]['5000']['F']
            m = gd_amp[age]['5000']['M']
            print(f"  edad {age:2d}: F={f:>5}  M={m:>5}")
        except KeyError as e:
            print(f"  edad {age}: FALTA {e}")


if __name__ == "__main__":
    main()
